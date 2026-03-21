"""
Tableau Export Generator
Produces formatted .xlsx and .csv files optimised for Tableau Desktop ingestion.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json

INPUT = Path("/home/claude/eu_china_tracker/data/analyzed_events.csv")
OUTPUT_DIR = Path("/home/claude/eu_china_tracker/tableau_exports")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

ACTIVITY_COLORS = {
    "Joint Research Projects":           "#1F77B4",
    "Joint Programs":                    "#FF7F0E",
    "Researcher Exchange":               "#2CA02C",
    "Conferences & Seminars":            "#D62728",
    "Exchange of Scientific Information":"#9467BD",
    "Flagship Initiatives":              "#8C564B",
    "Unclassified":                      "#BCBCBC",
}

def build_tableau_exports(df: pd.DataFrame):
    """Generate multiple Tableau-ready sheets."""

    # ── Sheet 1: Events master table ──────────────────────────────────────────
    master = df[[
        "year", "date", "title", "source", "language",
        "primary_activity", "secondary_activity",
        "classification_confidence", "sentiment_label", "sentiment_score",
        "entities_institutions", "entities_locations", "entities_domains",
        "url",
    ]].copy()
    master["activity_color"] = master["primary_activity"].map(ACTIVITY_COLORS)
    master["is_forecast"] = master["year"] >= 2026
    master["decade"] = (master["year"] // 10 * 10).astype(str) + "s"

    # ── Sheet 2: Activity counts by year ──────────────────────────────────────
    pivot_year = (
        df.groupby(["year", "primary_activity"])
        .size()
        .reset_index(name="event_count")
    )
    pivot_year["cumulative"] = pivot_year.groupby("primary_activity")["event_count"].cumsum()

    # ── Sheet 3: Sentiment by activity type ──────────────────────────────────
    sentiment = (
        df.groupby(["primary_activity", "sentiment_label"])
        .size()
        .reset_index(name="count")
    )
    sentiment_pivot = sentiment.pivot(
        index="primary_activity", columns="sentiment_label", values="count"
    ).fillna(0).reset_index()

    # ── Sheet 4: Language distribution ───────────────────────────────────────
    lang_year = (
        df.groupby(["year", "language"])
        .size()
        .reset_index(name="count")
    )

    # ── Sheet 5: Cross-activity matrix (co-occurrences) ───────────────────────
    co_occur = (
        df.groupby(["primary_activity", "secondary_activity"])
        .size()
        .reset_index(name="count")
        .dropna()
    )

    # ── Sheet 6: Trend summary (5-year rolling) ───────────────────────────────
    yearly_total = df.groupby("year").size().reset_index(name="total_events")
    yearly_total = yearly_total.sort_values("year")
    yearly_total["rolling_5yr"] = (
        yearly_total["total_events"].rolling(5, min_periods=1).mean().round(1)
    )
    yearly_total["yoy_change_pct"] = (
        yearly_total["total_events"].pct_change() * 100
    ).round(1)

    # ── Write to Excel (multi-sheet) ──────────────────────────────────────────
    xlsx_path = OUTPUT_DIR / "eu_china_st_tracker.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Events_Master", index=False)
        pivot_year.to_excel(writer, sheet_name="Activity_by_Year", index=False)
        sentiment_pivot.to_excel(writer, sheet_name="Sentiment_by_Activity", index=False)
        lang_year.to_excel(writer, sheet_name="Language_Distribution", index=False)
        co_occur.to_excel(writer, sheet_name="Activity_CoOccurrence", index=False)
        yearly_total.to_excel(writer, sheet_name="Trend_Summary", index=False)

        # Color-code the activity column in master sheet
        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets["Events_Master"]
        act_col_idx = list(master.columns).index("primary_activity") + 1
        color_col_idx = list(master.columns).index("activity_color") + 1
        for row_idx in range(2, len(master) + 2):
            cell = ws.cell(row=row_idx, column=act_col_idx)
            hex_color = ws.cell(row=row_idx, column=color_col_idx).value
            if hex_color:
                fill = PatternFill(
                    start_color=hex_color.lstrip("#"),
                    end_color=hex_color.lstrip("#"),
                    fill_type="solid"
                )
                cell.fill = fill
                cell.font = Font(color="FFFFFF", bold=True)

        # Format header rows
        header_fill = PatternFill(start_color="1a3d5c", end_color="1a3d5c", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for ws_name in writer.sheets:
            ws2 = writer.sheets[ws_name]
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws2.freeze_panes = "A2"

    print(f"✓ Excel export → {xlsx_path}")

    # ── Write individual CSVs for Tableau ────────────────────────────────────
    csv_files = {
        "events_master.csv": master,
        "activity_by_year.csv": pivot_year,
        "sentiment_by_activity.csv": sentiment,
        "language_distribution.csv": lang_year,
        "trend_summary.csv": yearly_total,
    }
    for fname, data in csv_files.items():
        p = OUTPUT_DIR / fname
        data.to_csv(p, index=False, encoding="utf-8-sig")
        print(f"✓ CSV → {p}")

    # ── Tableau TWB skeleton (XML) ─────────────────────────────────────────────
    twb_skeleton = """<?xml version='1.0' encoding='utf-8' ?>
<workbook source-build='2023.3' source-platform='win' version='21.4'>
  <datasources>
    <datasource caption='EU-China S&amp;T Tracker' name='eu_china_tracker'>
      <connection class='textscan' filename='events_master.csv' />
    </datasource>
  </datasources>
  <worksheets>
    <worksheet name='Activity Timeline' />
    <worksheet name='Stacked Bar by Year' />
    <worksheet name='Sentiment Heatmap' />
    <worksheet name='Language Breakdown' />
    <worksheet name='Trend Line' />
  </worksheets>
  <!-- 
    TABLEAU SETUP INSTRUCTIONS:
    1. Open Tableau Desktop and connect to events_master.csv
    2. Create calculated fields:
       - [Is Future] = [Year] >= 2026
       - [Decade] = STR(INT([Year]/10)*10) + 's'
    3. Recommended views:
       - Timeline: Year (cols) x COUNT(Events) (rows), Color = Primary Activity
       - Stacked: Year (cols) x Event Count (rows), Stack = Activity Type
       - Heatmap: Activity (rows) x Sentiment (cols), Size = Count
  -->
</workbook>"""
    twb_path = OUTPUT_DIR / "eu_china_tracker.twb"
    twb_path.write_text(twb_skeleton)
    print(f"✓ Tableau workbook skeleton → {twb_path}")

    return xlsx_path


if __name__ == "__main__":
    df = pd.read_csv(INPUT)
    xlsx = build_tableau_exports(df)
    print(f"\nAll Tableau exports ready in: {OUTPUT_DIR}")
